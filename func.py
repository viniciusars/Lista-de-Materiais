import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import Tk
from tkinter import filedialog
from pathlib import Path
from tkinter.filedialog import askopenfilename
from funcoes_acessorias import diretorio, local, escrever_csv

def pegar_postes():
    '''
    Lista que vai receber todos os as linhas válidas da tabela de locação
    Esses valores pré-definidos compõe o cabeçalho
    '''
    dados = [['Número','Vértice','Deflexão','Tipo','Altura/carga','Posição','Vão de frente','Distância progressiva','Nível 1','Circuito 1', 'Nível 2', 'Circuito 2', 'Âncora', 'Estais']]
    
    caminho = ""

    '''
    No momento que for para selecionar o caminho da tabela de locação pode ocorrer algum erro, dessa forma, para evitar 
    qualquer má funcionamento da função, deve-se verificar os seguintes pontos do arquivo selecionado:
    1 - Verificar se foi selecionado algum caminho, se não a função encerra
    2 - Verificar se é um arquivo .xlsx, caso isso não ocorra, a função é encerrada
    3 - Verificar se o arquivo .xlsx é uma tabela de locação, se não for, a função é encerrada

    Se tudo isso for confirmado, a função continua o funcionamento.
    '''

    while True:
        # Pegar o caminho do arquivo da tabela de locação que vai ser utilizada
        caminho = local('Selecione a tabela de locação')

        if not caminho:
            print('Não foi selecionado nenhum arquivo!!!')
            return print('Solicitação encerrada, chame novamente!!!')
        
        count = 0
        if caminho:
            if not '.xlsx' in caminho:
                print('O documento escolhido não é uma arquivo .xlsx')
                return print('Solicitação encerrada, chame novamente!!!') 
            
            wb_teste = load_workbook(caminho)
            for sheet in wb_teste.sheetnames:
                if 'TABELA DE LOCAÇÃO' in sheet:
                    count = count + 1

        if count > 0:
            break
        elif count == 0:
            print('Documento escolhido não é uma Tabela de Locação!!!')
            return print('Solicitação encerrada, chame novamente!!!')
            

    
    # Carrega o documento como um workbook pela biblioteca openpyxl
    wb = load_workbook(caminho)

    '''
    Esse FOR vai pegar todos os valores válidos para esse caso em particular.
    No documento selecionado, há outras abas que não possuem valores de interesse, dessa forma,
    só vamos localizar os valores que estão nas abas com o nome TABELA DE LOCAÇÃO 
    '''
    for sheets in wb.sheetnames: # Percorre o nome de todas as worksheets (abas da planilha)
        if "TABELA DE LOCAÇÃO" in sheets: # Verifica se, no nome da aba, possue o texto TABELA DE LOCAÇÃO (abas do nosso interesse)
            # Esse for vai iterar sobre todas as linhas das abas selecionadas
            for row in wb[sheets].iter_rows(min_row=8,  # Primeira linha dessa iteração
                                            max_row=82, # última linha 
                                            min_col=1,  # Primeira coluna
                                            max_col=14, # última coluna
                                            values_only=True # Pegar os valores de dentro da célula, se remover, ele pega a referência da célula
                                            ):
                    '''
                    O trecho a seguir verifica quantos elementos não possuem valores (None), se a linha toda for None,
                    essa linha não será adicionada a lista DADOS. Como 14 colunas são colhidas (14 valores), o for a seguir
                    itera sobre esses 14 valores verificando se são None, caso todos forem, essa linha não vai ser inserida
                    '''
                    valores_nulos = 0
                    for x, elemento in enumerate(row):
                        if elemento is None:
                            valores_nulos = valores_nulos + 1
                    if valores_nulos < 14:
                        dados.append(row)

    # No pensamento atual, faço a obtenção de todos esses valores e jogo para um arquivo .csv
    caminho_pasta = diretorio() + '/infos.csv' 
    # For para escrever todos os valores presentes na lista DADOS
    escrever_csv(caminho_pasta, dados)

def quant_postes():
    pd.set_option('future.no_silent_downcasting', True)
    '''
    Na função pegar_postes(), foi salvo o arquivo infos.csv que contém informações sobre os elementos presentes na tabela de locação.
    Dessa forma, é necessário pegar essa informação para fazer a manipulação e a contabilização desses materiais e é isso que será 
    realizado nessa função
    '''

    #! Parte para pegar o diretório do arquivo atual

    caminho = diretorio() + "/infos.csv" # Compatibiliza com o jeito correto
    postes = pd.read_csv(caminho, encoding= 'latin-1')
    '''
    O formato do arquivo tava dando conflito com o padrão da função, por isso, teve que fazer
    essa modificação
    '''
    
    postes_unicos = postes['Tipo'].value_counts() 
    '''
    O value_counts() conta quantas vezes um valor único aparece no df e, nesse caso, quantas vezes aparece na coluna.
    Isso é necessário para saber a quantidade de materiais x quantidade de cada estrutura 
    '''
    
    #! Mudança de caracteres
    '''
    A tabela de locação vem com '-', porém, as abas do excel não conseguem receber esses caracteres, por isso,
    nas planilhas das especificações dos postes vai com '_'. Com isso é preciso modificar a forma como é escrita 
    para compatibilizar
    '''
    for index in postes_unicos.index: # Os nomes incorretos estão no postes_unicos
        novo_index = index.replace('-','_')
        novo_index = novo_index.replace('/','_')
        postes_unicos = postes_unicos.rename(index={index : novo_index})

    #!  Aquisição da planilha com quais materiais vão em cada tipo de estrutura
    caminho_planilha_descricao_postes = diretorio() + '/POSTES.xlsx' 

    wb = load_workbook(caminho_planilha_descricao_postes)

    colunas = [] # Lista com nome de todas as colunas as quais vão ser o nome dos postes
    indice = [] # Lista com o index, nesse caso, vão ser todos os códigos de cada um dos materiais das estruturas

    for sheet in wb.sheetnames: # Percorre todas as abas da planilha POSTES que possue todas as montagens de todas as estruturas
        # Cada aba possui o nome de uma estrutura, por isso cada sheet vai ser adicionado
        for row in wb[sheet].iter_rows(min_row=2, max_col=2, min_col=2, values_only=True):
            indice.append(row) # Pegando o código de cada material
        colunas.append(sheet) # Pegando o nome da estrutura da coluna

    indice = pd.DataFrame(indice) 
    indice = indice[0].sort_values(ascending=True).unique()
    
    '''
    Como eu peguei os códigos de todas as abas, eles vão aparecer repetidos, dessa forma é necessário
    pagar só os valores únicos e colocar de forma crescente para deixar para organizado
    '''

    #! DF com a descrição por cada estrutura
    df_planilha_descricao_postes = pd.DataFrame(index= indice, columns=colunas) # Esse DF vai ter todos os materiais por cada estrutura

    '''
    O FOR seguinte vai inserir todos os valores de materiais por cada estrutura, a quantidade que cada material aparece 
    por estrutura 
    '''
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
            df_planilha_descricao_postes.loc[row[0] ,sheet] = row[2]

    #! Multiplicação de quantidade de materiais por estrutura x a quantidade de cada estrutura
    df_quantidade_total_materiais = pd.DataFrame(index= indice, columns=postes_unicos.index) # DF com o total de materiais por estrutura 

    # For para fazer a multiplicação
    for indices in df_planilha_descricao_postes.index:
        for estruturas in df_planilha_descricao_postes.columns:
            if estruturas in postes_unicos.index:
                '''
                A lista ESTRUTURAS aparecem todas os tipos de estruturas possíveis de existirem no parque, mas não quer dizer 
                que todas vão aparecer na tabela de locação de um parque, então, só serão escolhidas as estruturas que estão 
                em postes_unicos (lista que mostra todas as estruturas de um parque específico)
                '''
                if type(df_planilha_descricao_postes.loc[indices, estruturas]) is int: # Não selecionar os elementos que não possuem valores
                    df_quantidade_total_materiais.loc[indices, estruturas] = postes_unicos.loc[estruturas] * df_planilha_descricao_postes.loc[indices, estruturas]
                    # Feita a multiplicação entre a quantidade de elementos num tipo de estrutura x a quantidade de vezes que aquela estrutura aparece
    df_quantidade_total_materiais.fillna(0, inplace=True) # Remover os valores NaN
    
    df_quantidade_total_materiais['Valor Total'] = pd.DataFrame(0,index= indice, columns= ['Valor Total'])

    for indice in df_quantidade_total_materiais.index:
        for coluna in df_quantidade_total_materiais.columns:
            if coluna != 'Valor Total':
                df_quantidade_total_materiais.loc[indice, 'Valor Total'] = df_quantidade_total_materiais.loc[indice, 'Valor Total'] + df_quantidade_total_materiais.loc[indice, coluna] 

    df_quantidade_total_materiais.reset_index(inplace=True)
    df_quantidade_total_materiais.rename(columns={'index' :'Código do material'}, inplace=True)
    lista_quantidade_total_materiais = [df_quantidade_total_materiais.columns.tolist()] + df_quantidade_total_materiais.values.tolist()
    caminho_total_materiais = diretorio() + '/quantidade_materiais.csv'
    escrever_csv(caminho_total_materiais, lista_quantidade_total_materiais)

quant_postes()