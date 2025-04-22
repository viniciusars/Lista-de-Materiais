from funcoes_acessorias import local_de_salvar
import pandas as pd
from openpyxl import load_workbook


def quantidade_total():
    caminho = local_de_salvar('./data/Tabela_de_Locacao.csv')
    postes = pd.read_csv(caminho, encoding= 'latin-1')
    postes = postes['Tipo'].value_counts()
    postes = inserir_barra_nome(postes)
    return postes

def inserir_barra_nome(df):
    for index in df.index:
        novo_index = index.replace('/','|')
        df = df.rename(index={index: novo_index})
    return df

def preencher_tabela(worksheet, df):
    for sheet in worksheet.sheetnames:
        if sheet in df.columns:
            for row in worksheet[sheet].iter_rows(min_row=2, min_col=2, max_col=4, values_only=True):
                if type(row[2]) is int:
                    df.loc[row[0], sheet] = row[2]
                else:
                    df.loc[row[0], sheet] = 0
    return df

def multiplicacao_elementos_postes(df, serie):
    for estrutura in serie.index:
        for item in df.index:
            df.loc[item, estrutura] = df.loc[item, estrutura] * serie[estrutura]
    df['Valor Total'] = df.sum(axis=1)
    return df


def gerar_df_info_postes():
    wb_postes = load_workbook('./data/POSTES.xlsx')
    postes_unicos = quantidade_total()

    indices = []
    for sheet in wb_postes.sheetnames:
        for row in wb_postes[sheet].iter_rows(min_row=2, max_col=2,min_col=2, values_only=True):
            indices.append(row)
    indices = pd.DataFrame(indices)
    indices = indices[0].sort_values(ascending=True).unique()

    df_info_postes = pd.DataFrame(0, index= indices, columns= postes_unicos.index)
    df_info_postes = preencher_tabela(wb_postes, df_info_postes)
    df_info_postes = multiplicacao_elementos_postes(df_info_postes, postes_unicos)
    return df_info_postes

if __name__=="__main__":
    gerar_df_info_postes()

