from funcoes_acessorias import local, escrever_csv, local_de_salvar
from openpyxl import load_workbook

'''
Funcionamento desse módulo
Funções:
1 - verificar_arquivo() -           Esse primeira função vai servir para verificar se o arquivo escolhido
                                    foi escolhido corretamente. Verifica se foi escolhido algum arquivo,
                                    se foi escolhido um arquivo .xlsx e se, esse arquivo, é uma tabela de 
                                    locação.

2 - extrair_valores_planilha() -    Essa função vai pegar os valores que estão inseridos na tabela de 
                                    locação (sem formatação, sem um padrão fácil de manipular) e 
                                    transformar num arquivo .csv com colunas.

3 - pegar_postes() -                Função final. Vai utilizar algumas funções anteriores desse método
                                    e funções de outros métodos a fim de pegar a 'tabela de locação.xlsx'
                                    e transformar num arquivo .csv pronto para ser utilizado posteriormente
'''


def verificar_arquivo(caminho):
    if not caminho:
        print('Não foi selecionado nenhum arquivo')
        return False
    
    if not caminho.endswith('.xlsx'):
        print('Não é um arquivo .xlsx')
        return False
    
    wb = load_workbook(caminho)
    if not any('TABELA DE LOCAÇÃO' in sheet for sheet in wb.sheetnames):
        print('O arquivo não é de uma Tabela de Locação')
        return False
    return True



def extrair_valores_planilha(wb):
    dados: list = [['Número','Vértice','Deflexão','Tipo','Altura/carga','Posição','Vão de frente','Distância progressiva','Nível 1','Circuito 1', 'Nível 2', 'Circuito 2', 'Âncora', 'Estais']]
    for sheet in wb.sheetnames:
            if 'TABELA DE LOCAÇÃO' in sheet:
                for row in wb[sheet].iter_rows(min_row=8, max_row=82,min_col=1,max_col=14,values_only=True):
                    if sum(v is None for v in row) < 14:
                        dados.append(row)
    return dados



def pegar_postes():
    caminho: str = local('Arquivo da Tabela de Locação')
    verificador = verificar_arquivo(caminho) 

    if verificador:
        wb = load_workbook(caminho)
        dados = extrair_valores_planilha(wb)
        caminho = local_de_salvar('/data/Tabela_de_Locacao.csv')
        escrever_csv(caminho, dados)



if __name__ == '__main__':
    pegar_postes()

